import os
import csv
import io
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, verify_jwt_in_request
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None
from models import (
    ActivityLog,
    db,
    ClassGroup,
    Notification,
    Room,
    StudentStrength,
    Subject,
    Teacher,
    TeacherAvailability,
    TeacherSubject,
    TimeSlot,
    User,
    parse_time,
)
from utils.auth import role_required

data_bp = Blueprint("data", __name__, url_prefix="/api")


def _log_action(action: str):
    verify_jwt_in_request(optional=True)
    user_id = get_jwt_identity()
    db.session.add(ActivityLog(user_id=int(user_id) if user_id else None, action=action))


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ["1", "true", "yes", "y"]


def _read_rows_from_upload(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        text = file_storage.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if filename.endswith(".xlsx"):
        wb = load_workbook(file_storage, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for row in rows[1:]:
            out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return out

    if filename.endswith(".pdf"):
        if PdfReader is None:
            raise ValueError("PDF support requires pypdf. Run: pip install pypdf==5.2.0")
        reader = PdfReader(file_storage)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        if not lines:
            return []
        # Expected PDF text format: first line is comma-separated headers, remaining lines comma-separated values.
        headers = [h.strip() for h in lines[0].split(",")]
        out = []
        for line in lines[1:]:
            cols = [c.strip() for c in line.split(",")]
            out.append({headers[i]: cols[i] if i < len(cols) else None for i in range(len(headers))})
        return out

    raise ValueError("Unsupported file type. Use CSV, XLSX, or PDF.")


def _import_teachers(rows):
    created = 0
    for row in rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        max_per_day = int(row.get("max_lectures_per_day", 6) or 6)
        db.session.add(Teacher(name=name, max_lectures_per_day=max_per_day))
        created += 1
    return created


def _import_rooms(rows):
    created = 0
    for row in rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        capacity = int(row.get("capacity", 0) or 0)
        room_type = str(row.get("room_type", "classroom") or "classroom").strip().lower()
        db.session.add(Room(name=name, capacity=capacity, room_type=room_type))
        created += 1
    return created


def _import_classes(rows):
    created = 0
    existing = {c.name.strip().lower() for c in ClassGroup.query.all()}
    for row in rows:
        name = str(row.get("name", "")).strip()
        department = str(row.get("department", "")).strip()
        if not name or not department:
            continue
        if name.lower() in existing:
            continue
        strength = int(row.get("student_strength", 0) or 0)
        cls = ClassGroup(name=name, department=department, student_strength=strength)
        db.session.add(cls)
        db.session.flush()
        db.session.add(StudentStrength(class_id=cls.id, strength=strength))
        existing.add(name.lower())
        created += 1
    return created


def _import_subjects(rows):
    created = 0
    classes_by_name = {c.name.lower(): c.id for c in ClassGroup.query.all()}
    for row in rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        class_id = row.get("class_id")
        if not class_id:
            class_name = str(row.get("class_name", "")).strip().lower()
            class_id = classes_by_name.get(class_name)
        if not class_id:
            continue
        item = Subject(
            name=name,
            class_id=int(class_id),
            lectures_per_week=int(row.get("lectures_per_week", 1) or 1),
            priority_morning=_to_bool(row.get("priority_morning"), False),
            is_lab=_to_bool(row.get("is_lab"), False),
        )
        db.session.add(item)
        created += 1
    return created


def _import_users(rows):
    created = 0
    teachers_by_name = {t.name.lower(): t.id for t in Teacher.query.all()}
    classes_by_name = {c.name.lower(): c.id for c in ClassGroup.query.all()}
    for row in rows:
        username = str(row.get("username", "")).strip()
        password = str(row.get("password", "")).strip()
        role = str(row.get("role", "")).strip().lower()
        if not username or not password or role not in ["admin", "teacher", "student"]:
            continue
        if User.query.filter_by(username=username).first():
            continue
        teacher_id = row.get("teacher_id")
        class_id = row.get("class_id")
        if not teacher_id and row.get("teacher_name"):
            teacher_id = teachers_by_name.get(str(row.get("teacher_name")).strip().lower())
        if not class_id and row.get("class_name"):
            class_id = classes_by_name.get(str(row.get("class_name")).strip().lower())
        user = User(
            username=username,
            role=role,
            teacher_id=int(teacher_id) if teacher_id else None,
            class_id=int(class_id) if class_id else None,
        )
        user.set_password(password)
        db.session.add(user)
        created += 1
    return created


def _teacher_dict(t: Teacher):
    return {"id": t.id, "name": t.name, "max_lectures_per_day": t.max_lectures_per_day}


def _subject_dict(s: Subject):
    return {
        "id": s.id,
        "name": s.name,
        "class_id": s.class_id,
        "lectures_per_week": s.lectures_per_week,
        "priority_morning": s.priority_morning,
        "is_lab": s.is_lab,
    }


def _room_dict(r: Room):
    return {"id": r.id, "name": r.name, "capacity": r.capacity, "room_type": r.room_type}


def _class_dict(c: ClassGroup):
    return {"id": c.id, "name": c.name, "department": c.department, "student_strength": c.student_strength}


def _slot_dict(s: TimeSlot):
    return {
        "id": s.id,
        "day_of_week": s.day_of_week,
        "start_time": s.start_time.strftime("%H:%M"),
        "end_time": s.end_time.strftime("%H:%M"),
        "is_break": s.is_break,
        "slot_order": s.slot_order,
    }


@data_bp.get("/teachers")
@role_required("admin", "teacher", "student")
def get_teachers():
    return jsonify([_teacher_dict(t) for t in Teacher.query.all()])


@data_bp.post("/teachers")
@role_required("admin")
def create_teacher():
    data = request.get_json() or {}
    t = Teacher(name=data["name"], max_lectures_per_day=data.get("max_lectures_per_day", 6))
    db.session.add(t)
    _log_action(f"Created teacher {data['name']}")
    db.session.commit()
    return jsonify(_teacher_dict(t)), 201


@data_bp.put("/teachers/<int:teacher_id>")
@role_required("admin")
def update_teacher(teacher_id):
    t = Teacher.query.get_or_404(teacher_id)
    data = request.get_json() or {}
    t.name = data.get("name", t.name)
    t.max_lectures_per_day = data.get("max_lectures_per_day", t.max_lectures_per_day)
    _log_action(f"Updated teacher {teacher_id}")
    db.session.commit()
    return jsonify(_teacher_dict(t))


@data_bp.delete("/teachers/<int:teacher_id>")
@role_required("admin")
def delete_teacher(teacher_id):
    t = Teacher.query.get_or_404(teacher_id)
    db.session.delete(t)
    _log_action(f"Deleted teacher {teacher_id}")
    db.session.commit()
    return jsonify({"message": "Deleted"})


@data_bp.post("/teachers/<int:teacher_id>/subjects")
@role_required("admin")
def map_teacher_subject(teacher_id):
    data = request.get_json() or {}
    exists = TeacherSubject.query.filter_by(teacher_id=teacher_id, subject_id=data["subject_id"]).first()
    if exists:
        return jsonify({"message": "Mapping already exists", "id": exists.id, "teacher_id": exists.teacher_id, "subject_id": exists.subject_id}), 200
    rel = TeacherSubject(teacher_id=teacher_id, subject_id=data["subject_id"])
    db.session.add(rel)
    db.session.commit()
    return jsonify({"id": rel.id, "teacher_id": rel.teacher_id, "subject_id": rel.subject_id}), 201


@data_bp.get("/teacher-subjects")
@role_required("admin", "teacher")
def get_teacher_subjects():
    rows = TeacherSubject.query.all()
    return jsonify([{"id": r.id, "teacher_id": r.teacher_id, "subject_id": r.subject_id} for r in rows])


@data_bp.get("/subjects")
@role_required("admin", "teacher", "student")
def get_subjects():
    return jsonify([_subject_dict(s) for s in Subject.query.all()])


@data_bp.post("/subjects")
@role_required("admin")
def create_subject():
    data = request.get_json() or {}
    s = Subject(
        name=data["name"],
        class_id=data["class_id"],
        lectures_per_week=data["lectures_per_week"],
        priority_morning=data.get("priority_morning", False),
        is_lab=data.get("is_lab", False),
    )
    db.session.add(s)
    _log_action(f"Created subject {data['name']}")
    db.session.commit()
    return jsonify(_subject_dict(s)), 201


@data_bp.put("/subjects/<int:subject_id>")
@role_required("admin")
def update_subject(subject_id):
    s = Subject.query.get_or_404(subject_id)
    data = request.get_json() or {}
    s.name = data.get("name", s.name)
    s.class_id = data.get("class_id", s.class_id)
    s.lectures_per_week = data.get("lectures_per_week", s.lectures_per_week)
    s.priority_morning = data.get("priority_morning", s.priority_morning)
    s.is_lab = data.get("is_lab", s.is_lab)
    _log_action(f"Updated subject {subject_id}")
    db.session.commit()
    return jsonify(_subject_dict(s))


@data_bp.delete("/subjects/<int:subject_id>")
@role_required("admin")
def delete_subject(subject_id):
    s = Subject.query.get_or_404(subject_id)
    db.session.delete(s)
    _log_action(f"Deleted subject {subject_id}")
    db.session.commit()
    return jsonify({"message": "Deleted"})


@data_bp.get("/rooms")
@role_required("admin", "teacher", "student")
def get_rooms():
    return jsonify([_room_dict(r) for r in Room.query.all()])


@data_bp.post("/rooms")
@role_required("admin")
def create_room():
    data = request.get_json() or {}
    r = Room(name=data["name"], capacity=data["capacity"], room_type=data.get("room_type", "classroom"))
    db.session.add(r)
    _log_action(f"Created room {data['name']}")
    db.session.commit()
    return jsonify(_room_dict(r)), 201


@data_bp.put("/rooms/<int:room_id>")
@role_required("admin")
def update_room(room_id):
    r = Room.query.get_or_404(room_id)
    data = request.get_json() or {}
    r.name = data.get("name", r.name)
    r.capacity = data.get("capacity", r.capacity)
    r.room_type = data.get("room_type", r.room_type)
    _log_action(f"Updated room {room_id}")
    db.session.commit()
    return jsonify(_room_dict(r))


@data_bp.delete("/rooms/<int:room_id>")
@role_required("admin")
def delete_room(room_id):
    r = Room.query.get_or_404(room_id)
    db.session.delete(r)
    _log_action(f"Deleted room {room_id}")
    db.session.commit()
    return jsonify({"message": "Deleted"})


@data_bp.get("/classes")
@role_required("admin", "teacher", "student")
def get_classes():
    return jsonify([_class_dict(c) for c in ClassGroup.query.all()])


@data_bp.post("/classes")
@role_required("admin")
def create_class():
    data = request.get_json() or {}
    name = str(data.get("name", "")).strip()
    department = str(data.get("department", "")).strip()
    if not name or not department:
        return jsonify({"error": "name and department are required"}), 400
    if ClassGroup.query.filter_by(name=name).first():
        return jsonify({"error": f"Class '{name}' already exists"}), 409

    c = ClassGroup(name=data["name"], department=data["department"], student_strength=data.get("student_strength", 0))
    db.session.add(c)
    try:
        db.session.flush()
        db.session.add(StudentStrength(class_id=c.id, strength=c.student_strength))
        _log_action(f"Created class {data['name']}")
        db.session.commit()
        return jsonify(_class_dict(c)), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": f"Class '{name}' already exists"}), 409


@data_bp.put("/classes/<int:class_id>")
@role_required("admin")
def update_class(class_id):
    c = ClassGroup.query.get_or_404(class_id)
    data = request.get_json() or {}
    c.name = data.get("name", c.name)
    c.department = data.get("department", c.department)
    new_strength = data.get("student_strength", c.student_strength)
    if new_strength != c.student_strength:
        c.student_strength = new_strength
        db.session.add(StudentStrength(class_id=c.id, strength=new_strength))
    _log_action(f"Updated class {class_id}")
    db.session.commit()
    return jsonify(_class_dict(c))


@data_bp.delete("/classes/<int:class_id>")
@role_required("admin")
def delete_class(class_id):
    c = ClassGroup.query.get_or_404(class_id)
    db.session.delete(c)
    _log_action(f"Deleted class {class_id}")
    db.session.commit()
    return jsonify({"message": "Deleted"})


@data_bp.get("/timeslots")
@role_required("admin", "teacher", "student")
def get_timeslots():
    return jsonify([_slot_dict(s) for s in TimeSlot.query.order_by(TimeSlot.day_of_week, TimeSlot.slot_order).all()])


@data_bp.post("/timeslots")
@role_required("admin")
def create_timeslot():
    data = request.get_json() or {}
    s = TimeSlot(
        day_of_week=data["day_of_week"],
        start_time=parse_time(data["start_time"]),
        end_time=parse_time(data["end_time"]),
        is_break=data.get("is_break", False),
        slot_order=data["slot_order"],
    )
    db.session.add(s)
    db.session.commit()
    return jsonify(_slot_dict(s)), 201


@data_bp.delete("/timeslots/<int:slot_id>")
@role_required("admin")
def delete_timeslot(slot_id):
    s = TimeSlot.query.get_or_404(slot_id)
    db.session.delete(s)
    db.session.commit()
    return jsonify({"message": "Deleted"})


@data_bp.post("/availability")
@role_required("admin", "teacher")
def upsert_availability():
    data = request.get_json() or {}
    verify_jwt_in_request()
    user = User.query.get_or_404(int(get_jwt_identity()))
    if user.role == "teacher" and user.teacher_id != data["teacher_id"]:
        return jsonify({"error": "Teachers can only update their own availability"}), 403
    if user.role not in ["admin", "teacher"]:
        return jsonify({"error": "Forbidden"}), 403
    row = TeacherAvailability.query.filter_by(teacher_id=data["teacher_id"], time_slot_id=data["time_slot_id"]).first()
    if not row:
        row = TeacherAvailability(
            teacher_id=data["teacher_id"],
            time_slot_id=data["time_slot_id"],
            is_available=data.get("is_available", True),
        )
        db.session.add(row)
    else:
        row.is_available = data.get("is_available", row.is_available)
    _log_action(f"Updated availability for teacher {data['teacher_id']}")
    db.session.commit()
    return jsonify({"id": row.id, "teacher_id": row.teacher_id, "time_slot_id": row.time_slot_id, "is_available": row.is_available})


@data_bp.get("/availability/<int:teacher_id>")
@role_required("admin", "teacher")
def get_availability(teacher_id):
    rows = TeacherAvailability.query.filter_by(teacher_id=teacher_id).all()
    return jsonify([{"time_slot_id": r.time_slot_id, "is_available": r.is_available} for r in rows])


@data_bp.post("/users")
@role_required("admin")
def create_user():
    data = request.get_json() or {}
    u = User(username=data["username"], role=data["role"], teacher_id=data.get("teacher_id"), class_id=data.get("class_id"))
    u.set_password(data["password"])
    db.session.add(u)
    db.session.commit()
    return jsonify({"id": u.id, "username": u.username, "role": u.role}), 201


@data_bp.post("/settings/time-config")
@role_required("admin")
def configure_time_slots():
    """
    Build a fixed daily schedule for each working day:

    09:00 – 10:00  Slot 1 (1-hour opening theory)
    10:00 – 11:00  Slot 2
    11:00 – 12:00  Slot 3
    12:00 – 12:30  LUNCH BREAK (30 min) — always fixed
    12:30 – 13:30  Slot 4
    13:30 – 14:30  Slot 5
    14:30 – 14:45  SHORT BREAK (15 min) — 2 hours after lunch end
    14:45 – 15:45  Slot 6
    15:45 – 16:45  Slot 7
    """
    data = request.get_json() or {}
    working_days = data.get(
        "working_days",
        ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
    )

    # Fixed slot schedule (start_minutes, end_minutes, is_break, label)
    DAILY_SCHEDULE = [
        (9 * 60,       10 * 60,       False),   # Slot 1 — opening theory
        (10 * 60,      11 * 60,       False),   # Slot 2
        (11 * 60,      12 * 60,       False),   # Slot 3
        (12 * 60,      12 * 60 + 30,  True),    # LUNCH BREAK
        (12 * 60 + 30, 13 * 60 + 30,  False),   # Slot 4
        (13 * 60 + 30, 14 * 60 + 30,  False),   # Slot 5
        (14 * 60 + 30, 14 * 60 + 45,  True),    # SHORT BREAK (2 h after lunch)
        (14 * 60 + 45, 15 * 60 + 45,  False),   # Slot 6
        (15 * 60 + 45, 16 * 60 + 45,  False),   # Slot 7
    ]

    def _fmt(minutes):
        return f"{minutes // 60:02d}:{minutes % 60:02d}"

    TimeSlot.query.delete()
    for day in working_days:
        for order, (start, end, is_brk) in enumerate(DAILY_SCHEDULE, start=1):
            s = TimeSlot(
                day_of_week=day,
                start_time=parse_time(_fmt(start)),
                end_time=parse_time(_fmt(end)),
                is_break=is_brk,
                slot_order=order,
            )
            db.session.add(s)

    _log_action("Configured global time slots (fixed schedule)")
    db.session.commit()
    return jsonify({"message": "Time slots configured with fixed break schedule"})



@data_bp.get("/student-strength/<int:class_id>")
@role_required("admin")
def student_strength_history(class_id):
    rows = StudentStrength.query.filter_by(class_id=class_id).order_by(StudentStrength.updated_at.desc()).all()
    return jsonify([{"strength": r.strength, "updated_at": r.updated_at.isoformat()} for r in rows])


@data_bp.get("/notifications")
@role_required("admin", "teacher", "student")
def get_notifications():
    verify_jwt_in_request()
    user_id = int(get_jwt_identity())
    rows = Notification.query.filter_by(user_id=user_id).order_by(Notification.created_at.desc()).limit(30).all()
    return jsonify([{"id": n.id, "title": n.title, "message": n.message, "is_read": n.is_read, "created_at": n.created_at.isoformat()} for n in rows])


@data_bp.post("/notifications/<int:notification_id>/read")
@role_required("admin", "teacher", "student")
def read_notification(notification_id):
    verify_jwt_in_request()
    user_id = int(get_jwt_identity())
    row = Notification.query.get_or_404(notification_id)
    if row.user_id != user_id:
        return jsonify({"error": "Forbidden"}), 403
    row.is_read = True
    db.session.commit()
    return jsonify({"message": "Marked as read"})


@data_bp.get("/backup/export")
@role_required("admin")
def export_backup():
    db_path = db.engine.url.database
    if not db_path or not os.path.exists(db_path):
        return jsonify({"error": "Database file not found"}), 404
    return send_file(db_path, as_attachment=True, download_name="smart_timetable_backup.sqlite")


@data_bp.post("/backup/restore")
@role_required("admin")
def restore_backup():
    if "file" not in request.files:
        return jsonify({"error": "Upload a SQLite backup file"}), 400
    db_path = db.engine.url.database
    backup_file = request.files["file"]
    if not backup_file.filename.endswith((".sqlite", ".db")):
        return jsonify({"error": "Invalid file type"}), 400
    db.session.remove()
    backup_file.save(db_path)
    return jsonify({"message": "Backup restored. Restart backend to reconnect safely."})


@data_bp.get("/activity-logs")
@role_required("admin")
def get_activity_logs():
    rows = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(200).all()
    return jsonify([{"id": r.id, "user_id": r.user_id, "action": r.action, "created_at": r.created_at.isoformat()} for r in rows])


@data_bp.post("/bulk/upload")
@role_required("admin")
def bulk_upload():
    target = (request.form.get("target") or "").strip().lower()
    if target not in ["teachers", "rooms", "classes", "subjects", "users"]:
        return jsonify({"error": "Invalid target. Use teachers, rooms, classes, subjects, or users."}), 400
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    uploaded = request.files["file"]
    if not uploaded.filename:
        return jsonify({"error": "File name is missing"}), 400

    try:
        rows = _read_rows_from_upload(uploaded)
    except ValueError as ex:
        return jsonify({"error": str(ex)}), 400
    except Exception:
        return jsonify({"error": "Unable to read uploaded file"}), 400

    if not rows:
        return jsonify({"error": "No rows found in file"}), 400

    try:
        created = 0
        if target == "teachers":
            created = _import_teachers(rows)
        elif target == "rooms":
            created = _import_rooms(rows)
        elif target == "classes":
            created = _import_classes(rows)
        elif target == "subjects":
            created = _import_subjects(rows)
        elif target == "users":
            created = _import_users(rows)
        db.session.flush()
        _log_action(f"Bulk uploaded {created} records into {target}")
        db.session.commit()
        return jsonify({"message": "Bulk upload complete", "target": target, "rows_read": len(rows), "rows_imported": created})
    except Exception as ex:
        db.session.rollback()
        return jsonify({"error": f"Bulk upload failed: {str(ex)}"}), 400


@data_bp.get("/bulk/template/<string:target>")
@role_required("admin")
def bulk_template(target):
    target = target.strip().lower()
    templates = {
        "teachers": "name,max_lectures_per_day\nJohn Doe,5\nJane Smith,6\n",
        "rooms": "name,capacity,room_type\nA101,60,classroom\nLab-1,40,lab\n",
        "classes": "name,department,student_strength\nBSc-CS-1,Computer Science,55\nMBA-1,Management,48\n",
        "subjects": "name,class_name,lectures_per_week,priority_morning,is_lab\nMathematics,BSc-CS-1,4,true,false\nDBMS Lab,BSc-CS-1,2,false,true\n",
        "users": "username,password,role,teacher_name,class_name\nteacher_john,secret123,teacher,John Doe,\nstudent_001,secret123,student,,BSc-CS-1\n",
    }
    if target not in templates:
        return jsonify({"error": "Invalid target"}), 400
    return templates[target], 200, {"Content-Type": "text/csv; charset=utf-8", "Content-Disposition": f"attachment; filename={target}_template.csv"}
