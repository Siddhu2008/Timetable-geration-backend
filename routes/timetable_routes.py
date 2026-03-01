from collections import defaultdict
from io import BytesIO
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, verify_jwt_in_request
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from models import (
    db,
    ClassGroup,
    ConflictLog,
    Notification,
    Room,
    ScheduleChangeRequest,
    Subject,
    Teacher,
    TeacherAvailability,
    TeacherSubject,
    TimeSlot,
    TimetableEntry,
    TimetableVersion,
    User,
)
from services.timetable_service import detect_conflicts, generate_timetable
from utils.auth import role_required

timetable_bp = Blueprint("timetable", __name__, url_prefix="/api/timetable")


def _entry_dict(e):
    return {
        "id": e.id,
        "version_id": e.version_id,
        "class_id": e.class_id,
        "subject_id": e.subject_id,
        "teacher_id": e.teacher_id,
        "room_id": e.room_id,
        "time_slot_id": e.time_slot_id,
        "is_locked": e.is_locked,
    }


def _is_teacher_available(teacher_id: int, slot_id: int) -> bool:
    row = TeacherAvailability.query.filter_by(teacher_id=teacher_id, time_slot_id=slot_id).first()
    return True if not row else bool(row.is_available)


def _validate_entry(version_id: int, entry_id: int, class_id: int, teacher_id: int, room_id: int, slot_id: int):
    target_slot = TimeSlot.query.get(slot_id)
    if not target_slot or target_slot.is_break:
        return "Cannot assign timetable entry to a break slot."
    if not _is_teacher_available(teacher_id, slot_id):
        return "Teacher is unavailable in the target slot."

    cls = ClassGroup.query.get(class_id)
    room = Room.query.get(room_id)
    if cls and room and room.capacity < cls.student_strength:
        return "Room capacity does not match class strength."

    same_slot_entries = TimetableEntry.query.filter_by(version_id=version_id, time_slot_id=slot_id).all()
    for row in same_slot_entries:
        if row.id == entry_id:
            continue
        if row.teacher_id == teacher_id:
            return "Teacher clash detected."
        if row.room_id == room_id:
            return "Room clash detected."
        if row.class_id == class_id:
            return "Class overlap detected."

    subject_day_dup = (
        db.session.query(TimetableEntry)
        .join(TimeSlot, TimetableEntry.time_slot_id == TimeSlot.id)
        .filter(
            TimetableEntry.version_id == version_id,
            TimetableEntry.id != entry_id,
            TimetableEntry.class_id == class_id,
            TimetableEntry.subject_id == TimetableEntry.query.get(entry_id).subject_id,
            TimeSlot.day_of_week == target_slot.day_of_week,
        )
        .first()
    )
    if subject_day_dup:
        return "Same subject repetition in one day is not allowed."
    return None


def _suggest_alternate_slots(version_id: int, entry: TimetableEntry, limit: int = 5):
    suggestions = []
    all_slots = TimeSlot.query.filter_by(is_break=False).order_by(TimeSlot.day_of_week, TimeSlot.slot_order).all()
    for slot in all_slots:
        err = _validate_entry(version_id, entry.id, entry.class_id, entry.teacher_id, entry.room_id, slot.id)
        if not err:
            suggestions.append({"time_slot_id": slot.id, "day": slot.day_of_week, "start": slot.start_time.strftime("%H:%M"), "end": slot.end_time.strftime("%H:%M")})
        if len(suggestions) >= limit:
            break
    return suggestions


@timetable_bp.post("/generate")
@role_required("admin")
def generate():
    verify_jwt_in_request()
    data = request.get_json() or {}
    classes = ClassGroup.query.all()
    subjects = Subject.query.all()
    subject_ids_with_teachers = {r.subject_id for r in TeacherSubject.query.all()}
    issues = []
    for cls in classes:
        cls_subjects = [s for s in subjects if s.class_id == cls.id]
        if not cls_subjects:
            issues.append(f"Class '{cls.name}' has no subjects configured.")
            continue
        for sub in cls_subjects:
            if sub.id not in subject_ids_with_teachers:
                issues.append(f"Subject '{sub.name}' in class '{cls.name}' has no teacher mapping.")
    if issues:
        return jsonify({"error": "Generation blocked due to incomplete class-wise setup.", "issues": issues}), 400

    versions = generate_timetable(
        created_by=int(get_jwt_identity()),
        num_versions=data.get("num_versions", 3),
        max_retries=data.get("max_retries", 80),
    )
    if not versions:
        return jsonify({"error": "Generation failed. Add more constraints/resources."}), 400
    return jsonify([{"id": v.id, "name": v.name, "score": v.score, "is_active": v.is_active} for v in versions]), 201


@timetable_bp.get("/versions")
@role_required("admin", "teacher", "student")
def versions():
    rows = TimetableVersion.query.order_by(TimetableVersion.created_at.desc()).all()
    return jsonify([{"id": v.id, "name": v.name, "score": v.score, "is_active": v.is_active} for v in rows])


@timetable_bp.post("/versions/<int:version_id>/activate")
@role_required("admin")
def activate(version_id):
    TimetableVersion.query.update({"is_active": False})
    row = TimetableVersion.query.get_or_404(version_id)
    row.is_active = True
    db.session.commit()
    return jsonify({"message": "Activated"})


@timetable_bp.get("/preview/<int:version_id>")
@role_required("admin", "teacher", "student")
def preview(version_id):
    entries = TimetableEntry.query.filter_by(version_id=version_id).all()
    return jsonify([_entry_dict(e) for e in entries])


@timetable_bp.post("/lock")
@role_required("admin")
def lock_period():
    data = request.get_json() or {}
    e = TimetableEntry(
        version_id=data.get("version_id", 0),
        class_id=data["class_id"],
        subject_id=data["subject_id"],
        teacher_id=data["teacher_id"],
        room_id=data["room_id"],
        time_slot_id=data["time_slot_id"],
        is_locked=True,
    )
    db.session.add(e)
    db.session.commit()
    return jsonify(_entry_dict(e)), 201


@timetable_bp.get("/conflicts/<int:version_id>")
@role_required("admin")
def conflicts(version_id):
    detected = detect_conflicts(version_id)
    stored = ConflictLog.query.filter_by(version_id=version_id).all()
    suggestions = []
    entries = TimetableEntry.query.filter_by(version_id=version_id).all()
    for entry in entries[:25]:
        alt = _suggest_alternate_slots(version_id, entry, limit=2)
        if alt:
            suggestions.append({"entry_id": entry.id, "class_id": entry.class_id, "subject_id": entry.subject_id, "alternates": alt})
    return jsonify(
        {
            "detected": detected,
            "history": [{"type": c.conflict_type, "message": c.message, "created_at": c.created_at.isoformat()} for c in stored],
            "suggested_alternates": suggestions,
        }
    )


@timetable_bp.post("/entry/<int:entry_id>/move")
@role_required("admin")
def move_entry(entry_id):
    data = request.get_json() or {}
    entry = TimetableEntry.query.get_or_404(entry_id)
    if entry.is_locked:
        return jsonify({"error": "Locked entry cannot be moved"}), 400

    new_slot = data.get("time_slot_id")
    if not new_slot:
        return jsonify({"error": "time_slot_id is required"}), 400

    issue = _validate_entry(entry.version_id, entry.id, entry.class_id, entry.teacher_id, entry.room_id, int(new_slot))
    if issue:
        return jsonify({"error": issue, "alternates": _suggest_alternate_slots(entry.version_id, entry)}), 400

    entry.time_slot_id = int(new_slot)
    db.session.commit()
    return jsonify({"message": "Entry moved successfully", "entry": _entry_dict(entry)})


@timetable_bp.get("/teacher")
@role_required("teacher")
def teacher_view():
    verify_jwt_in_request()
    user = User.query.get(int(get_jwt_identity()))
    active = TimetableVersion.query.filter_by(is_active=True).first()
    if not active:
        return jsonify([])
    rows = TimetableEntry.query.filter_by(version_id=active.id, teacher_id=user.teacher_id).all()
    return jsonify([_entry_dict(r) for r in rows])


@timetable_bp.post("/teacher/request-change")
@role_required("teacher")
def teacher_request_change():
    verify_jwt_in_request()
    data = request.get_json() or {}
    user = User.query.get(int(get_jwt_identity()))
    entry = TimetableEntry.query.get_or_404(data["timetable_entry_id"])
    if entry.teacher_id != user.teacher_id:
        return jsonify({"error": "You can request changes only for your own timetable entries"}), 403
    req = ScheduleChangeRequest(teacher_id=user.teacher_id, timetable_entry_id=entry.id, reason=data.get("reason", "No reason provided"))
    db.session.add(req)
    admins = User.query.filter_by(role="admin").all()
    for admin in admins:
        db.session.add(
            Notification(
                user_id=admin.id,
                title="Schedule Change Request",
                message=f"Teacher {user.username} requested a schedule change for entry #{entry.id}",
            )
        )
    db.session.commit()
    return jsonify({"message": "Request submitted"}), 201


@timetable_bp.get("/teacher/change-requests")
@role_required("admin", "teacher")
def list_change_requests():
    verify_jwt_in_request()
    user = User.query.get(int(get_jwt_identity()))
    q = ScheduleChangeRequest.query
    if user.role == "teacher":
        q = q.filter_by(teacher_id=user.teacher_id)
    rows = q.order_by(ScheduleChangeRequest.created_at.desc()).all()
    return jsonify(
        [{"id": r.id, "teacher_id": r.teacher_id, "timetable_entry_id": r.timetable_entry_id, "reason": r.reason, "status": r.status, "created_at": r.created_at.isoformat()} for r in rows]
    )


@timetable_bp.post("/teacher/change-requests/<int:request_id>/status")
@role_required("admin")
def update_change_request_status(request_id):
    data = request.get_json() or {}
    row = ScheduleChangeRequest.query.get_or_404(request_id)
    row.status = data.get("status", row.status)
    db.session.commit()
    return jsonify({"message": "Status updated"})


@timetable_bp.get("/student")
@role_required("student")
def student_view():
    verify_jwt_in_request()
    user = User.query.get(int(get_jwt_identity()))
    active = TimetableVersion.query.filter_by(is_active=True).first()
    if not active:
        return jsonify([])
    rows = TimetableEntry.query.filter_by(version_id=active.id, class_id=user.class_id).all()
    return jsonify([_entry_dict(r) for r in rows])


def _report_payload(version_id: int):
    entries = TimetableEntry.query.filter_by(version_id=version_id).all()
    slots = {s.id: s for s in TimeSlot.query.all()}
    subjects = {s.id: s for s in Subject.query.all()}
    teachers = {t.id: t for t in Teacher.query.all()}
    classes = {c.id: c for c in ClassGroup.query.all()}

    teacher_workload = defaultdict(int)
    room_usage = defaultdict(int)
    subject_distribution = defaultdict(int)
    free_slots = defaultdict(int)

    class_day_slot = defaultdict(set)
    for e in entries:
        teacher_workload[teachers[e.teacher_id].name] += 1
        room_usage[str(e.room_id)] += 1
        subject_distribution[subjects[e.subject_id].name] += 1
        class_day_slot[(classes[e.class_id].name, slots[e.time_slot_id].day_of_week)].add(e.time_slot_id)

    all_slots_per_day = defaultdict(list)
    for s in TimeSlot.query.filter_by(is_break=False).all():
        all_slots_per_day[s.day_of_week].append(s.id)

    for cls in classes.values():
        for day, ids in all_slots_per_day.items():
            used = len(class_day_slot[(cls.name, day)])
            free_slots[f"{cls.name}-{day}"] = max(0, len(ids) - used)

    version = TimetableVersion.query.get(version_id)
    conflict_count = len(detect_conflicts(version_id))
    return {
        "teacher_workload": teacher_workload,
        "room_usage": room_usage,
        "subject_distribution": subject_distribution,
        "free_slot_analysis": free_slots,
        "optimization_score": version.score if version else 0,
        "conflict_count": conflict_count,
    }


@timetable_bp.get("/reports/<int:version_id>")
@role_required("admin")
def reports(version_id):
    return jsonify(_report_payload(version_id))


@timetable_bp.get("/reports/export/excel/<int:version_id>")
@role_required("admin")
def export_report_excel(version_id):
    payload = _report_payload(version_id)
    wb = Workbook()
    for sheet_name, data in payload.items():
        if not isinstance(data, dict):
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(["metric", "value"])
        for key, value in data.items():
            ws.append([str(key), value])
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"reports_{version_id}.xlsx")


@timetable_bp.get("/reports/export/pdf/<int:version_id>")
@role_required("admin")
def export_report_pdf(version_id):
    payload = _report_payload(version_id)
    output = BytesIO()
    p = canvas.Canvas(output, pagesize=A4)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(40, 810, f"Report Summary - Version {version_id}")
    y = 780
    for section, data in payload.items():
        p.setFont("Helvetica-Bold", 10)
        p.drawString(40, y, section.replace("_", " ").title())
        y -= 18
        p.setFont("Helvetica", 9)
        if isinstance(data, dict):
            for k, v in list(data.items())[:14]:
                p.drawString(50, y, f"{k}: {v}")
                y -= 14
                if y < 60:
                    p.showPage()
                    y = 800
        else:
            p.drawString(50, y, str(data))
            y -= 18
    p.save()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"reports_{version_id}.pdf")


@timetable_bp.get("/export/excel/<int:version_id>")
@role_required("admin", "teacher", "student")
def export_excel(version_id):
    rows = TimetableEntry.query.filter_by(version_id=version_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    headers = ["id", "version_id", "class_id", "subject_id", "teacher_id", "room_id", "time_slot_id", "is_locked"]
    ws.append(headers)
    for r in rows:
        ws.append([r.id, r.version_id, r.class_id, r.subject_id, r.teacher_id, r.room_id, r.time_slot_id, r.is_locked])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"timetable_{version_id}.xlsx")


@timetable_bp.get("/export/pdf/<int:version_id>")
@role_required("admin", "teacher", "student")
def export_pdf(version_id):
    rows = TimetableEntry.query.filter_by(version_id=version_id).all()
    output = BytesIO()
    p = canvas.Canvas(output, pagesize=landscape(A4))
    p.setFont("Helvetica", 11)
    p.drawString(30, 570, f"Timetable Version {version_id}")
    y = 540
    for r in rows[:24]:
        p.drawString(30, y, f"Class {r.class_id} | Subject {r.subject_id} | Teacher {r.teacher_id} | Room {r.room_id} | Slot {r.time_slot_id}")
        y -= 20
        if y < 40:
            p.showPage()
            y = 540
    p.save()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"timetable_{version_id}.pdf")


@timetable_bp.get("/substitutes/needed")
@role_required("admin")
def substitutes_needed():
    active = TimetableVersion.query.filter_by(is_active=True).first()
    if not active:
        return jsonify([])

    # Find all entries where the assigned teacher is marked unavailable for that slot
    entries = TimetableEntry.query.filter_by(version_id=active.id).all()
    unavailabilities = TeacherAvailability.query.filter_by(is_available=False).all()
    unavailable_set = {(u.teacher_id, u.time_slot_id) for u in unavailabilities}

    needed = []
    teachers = {t.id: t for t in Teacher.query.all()}
    classes = {c.id: c for c in ClassGroup.query.all()}
    subjects = {s.id: s for s in Subject.query.all()}
    slots = {s.id: s for s in TimeSlot.query.all()}

    # Calculate free teachers for each slot
    all_teachers = set(teachers.keys())
    slot_to_busy_teachers = defaultdict(set)
    for e in entries:
        slot_to_busy_teachers[e.time_slot_id].add(e.teacher_id)
        
    for u in unavailabilities:
        slot_to_busy_teachers[u.time_slot_id].add(u.teacher_id)

    for e in entries:
        if (e.teacher_id, e.time_slot_id) in unavailable_set:
            absent_teacher = teachers.get(e.teacher_id)
            cls = classes.get(e.class_id)
            subj = subjects.get(e.subject_id)
            slot = slots.get(e.time_slot_id)

            available_teacher_ids = all_teachers - slot_to_busy_teachers[e.time_slot_id]
            available_teachers_list = [{"id": tid, "name": teachers[tid].name} for tid in available_teacher_ids]

            needed.append({
                "entry_id": e.id,
                "absent_teacher_id": e.teacher_id,
                "absent_teacher_name": absent_teacher.name if absent_teacher else "Unknown",
                "class_name": cls.name if cls else "Unknown",
                "subject_name": subj.name if subj else "Unknown",
                "day": slot.day_of_week if slot else "Unknown",
                "time": f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}" if slot else "Unknown",
                "available_substitutes": available_teachers_list
            })

    return jsonify(needed)


@timetable_bp.post("/substitutes/assign")
@role_required("admin")
def assign_substitute():
    data = request.get_json() or {}
    entry_id = data.get("entry_id")
    new_teacher_id = data.get("substitute_teacher_id")

    if not entry_id or not new_teacher_id:
        return jsonify({"error": "entry_id and substitute_teacher_id are required"}), 400

    entry = TimetableEntry.query.get_or_404(entry_id)
    entry.teacher_id = new_teacher_id
    db.session.commit()

    return jsonify({"message": "Substitute assigned successfully"})
